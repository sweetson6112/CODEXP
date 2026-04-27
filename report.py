"""
Excel report generation for CDRSL Barcode & Expiry Report.
Uses openpyxl to produce a styled, print-ready workbook.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─── Color Palette ────────────────────────────────────────────────────────────
DARK_BLUE  = "1a237e"
MID_BLUE   = "1565c0"
LIGHT_BLUE = "e3f2fd"
ORANGE_BG  = "fff3e0"
GREEN_BG   = "e8f5e9"
WHITE      = "FFFFFF"
GREY_HDR   = "f5f5f5"
WARN_RED   = "ffebee"


def _thin_border():
    side = Side(style="thin", color="bdbdbd")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _cell_font(size=10, bold=False, color="212121"):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def generate_report_excel(header: dict, lines: list) -> bytes:
    """
    Generate an Excel report for a barcode/expiry inward session.

    Args:
        header: dict from inward_headers table
        lines:  list of dicts from inward_line_items table

    Returns:
        bytes of the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Barcode & Expiry Report"
    ws.sheet_view.showGridLines = False

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [6, 14, 30, 20, 16, 16, 14, 10, 22, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ══════════════════════════════════════════════════════════════════════════
    # TITLE BLOCK
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{row}:J{row}")
    title_cell = ws[f"A{row}"]
    title_cell.value = "CDRSL — BARCODE & EXPIRY REPORT"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(f"A{row}:J{row}")
    sub_cell = ws[f"A{row}"]
    sub_cell.value = "Inward Inspection · Barcode Verification · Shelf Life Tracking"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color=WHITE)
    sub_cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # HEADER DETAILS BLOCK
    # ══════════════════════════════════════════════════════════════════════════
    def _hdr_label(ws, r, c, text):
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = Font(name="Calibri", size=9, bold=True, color=DARK_BLUE)
        cell.fill = PatternFill("solid", fgColor="e8eaf6")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def _hdr_value(ws, r, c, text, merge_end=None):
        if merge_end:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_end)
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = Font(name="Calibri", size=10, bold=True, color="0d47a1")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    hdr_section_start = row
    # Row 1 of header
    ws.row_dimensions[row].height = 20
    _hdr_label(ws, row, 1, "Document No")
    _hdr_value(ws, row, 2, header.get("document_no",""), merge_end=3)
    _hdr_label(ws, row, 4, "Date & Time")
    _hdr_value(ws, row, 5, str(header.get("session_datetime",""))[:19], merge_end=6)
    _hdr_label(ws, row, 7, "Inward Date")
    _hdr_value(ws, row, 8, header.get("inward_date",""), merge_end=9)
    _hdr_label(ws, row, 10, "Exp Threshold %")
    row += 1

    ws.row_dimensions[row].height = 20
    _hdr_label(ws, row, 1, "Invoice No")
    _hdr_value(ws, row, 2, header.get("invoice_no",""), merge_end=3)
    _hdr_label(ws, row, 4, "File No")
    _hdr_value(ws, row, 5, header.get("file_no",""), merge_end=6)
    _hdr_label(ws, row, 7, "Bill of Entry No")
    _hdr_value(ws, row, 8, header.get("boe_no",""), merge_end=9)
    ws.cell(row=row-1, column=10).value = f"{header.get('expiry_threshold',0)}%"
    ws.cell(row=row-1, column=10).font = Font(name="Calibri", size=11, bold=True, color="b71c1c")
    ws.cell(row=row-1, column=10).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(row=row-1, column=10).border = _thin_border()
    ws.cell(row=row-1, column=10).alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws.row_dimensions[row].height = 20
    _hdr_label(ws, row, 1, "BOE Date")
    _hdr_value(ws, row, 2, header.get("boe_date",""), merge_end=3)
    _hdr_label(ws, row, 4, "Container No")
    _hdr_value(ws, row, 5, header.get("container_no",""), merge_end=6)
    _hdr_label(ws, row, 7, "Status")
    _hdr_value(ws, row, 8, header.get("status",""), merge_end=9)
    row += 1

    ws.row_dimensions[row].height = 20
    _hdr_label(ws, row, 1, "Goods Receipt Date")
    _hdr_value(ws, row, 2, header.get("goods_receipt_date",""), merge_end=3)
    _hdr_label(ws, row, 4, "Invoice Lines")
    _hdr_value(ws, row, 5, header.get("invoice_lines",""), merge_end=6)
    _hdr_label(ws, row, 7, "Actual Lines")
    _hdr_value(ws, row, 8, header.get("actual_lines",""), merge_end=9)
    row += 2

    # Verified by
    ws.merge_cells(f"A{row}:J{row}")
    vb_cell = ws[f"A{row}"]
    vb_cell.value = f"Barcode/Expiry Verification done by:  {header.get('verified_by','')}"
    vb_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    vb_cell.fill = PatternFill("solid", fgColor="e8eaf6")
    vb_cell.alignment = Alignment(horizontal="left", vertical="center")
    vb_cell.border = _thin_border()
    ws.row_dimensions[row].height = 22
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # LINE ITEMS SECTION
    # ══════════════════════════════════════════════════════════════════════════
    line_cols = [
        "Sl.No", "Item No", "Item Description", "Barcode",
        "Expiry Date", "Mfg Date", "Shelf Life %", "Qty", "Remarks", "Verified By"
    ]
    ws.row_dimensions[row].height = 24
    for ci, col_name in enumerate(line_cols, start=1):
        cell = ws.cell(row=row, column=ci, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    row += 1

    threshold = float(header.get("expiry_threshold", 70))

    for idx, line in enumerate(lines, start=1):
        ws.row_dimensions[row].height = 18
        shelf_pct = line.get("shelf_life_pct")
        below = (shelf_pct is not None and shelf_pct < threshold)
        row_fill = PatternFill("solid", fgColor=(WARN_RED if below else (GREY_HDR if idx % 2 == 0 else WHITE)))

        values = [
            idx,
            line.get("item_no",""),
            line.get("description",""),
            line.get("barcode",""),
            line.get("expiry_date",""),
            line.get("mfg_date",""),
            f"{shelf_pct}%" if shelf_pct is not None else "—",
            line.get("qty",""),
            line.get("remark",""),
            line.get("verified_by",""),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _cell_font()
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center" if ci in [1,7,8] else "left",
                                       vertical="center", wrap_text=True)

            # Colour shelf life cell
            if ci == 7:
                if below:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="c62828")
                else:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="1b5e20")

            # Colour remarks
            if ci == 9:
                if line.get("remark") == "New Item":
                    cell.font = Font(name="Calibri", size=10, bold=True, color="1b5e20")
                elif line.get("remark") == "Already Barcode Exists":
                    cell.font = Font(name="Calibri", size=10, bold=True, color="e65100")

        row += 1

    # ── Summary row ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value=f"Total Items: {len(lines)}").font = Font(bold=True, color=DARK_BLUE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ── Footer ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:J{row}")
    footer = ws[f"A{row}"]
    footer.value = f"Created by Sweetson Joseph  |  Generated on {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}  |  CDRSL Barcode & Expiry System"
    footer.font = Font(name="Calibri", size=8, italic=True, color="757575")
    footer.alignment = Alignment(horizontal="center")

    # ── Freeze panes at line items table ───────────────────────────────────
    ws.freeze_panes = ws.cell(row=hdr_section_start + 8, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
